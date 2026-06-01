"""XLSX parser using openpyxl.

Migrated from LightRAG native XLSX parser.  Extracts text from all
sheets, using tab-delimited format for cells and sheet separators.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from aurora_ext.rag.parser.base import BaseParser, ParseResult

logger = logging.getLogger(__name__)


class XlsxParser(BaseParser):
    """Parser for .xlsx files using ``openpyxl``."""

    _EXTENSIONS = {"xlsx", "xls"}

    @property
    def supported_extensions(self) -> set[str]:
        return self._EXTENSIONS

    async def parse(self, file_path: str | Path, **kwargs: Any) -> ParseResult:
        from openpyxl import load_workbook

        path = Path(file_path)

        try:
            wb = load_workbook(str(path), read_only=True, data_only=True)
        except Exception as exc:
            raise ValueError(f"Cannot open XLSX: {path}") from exc

        sheet_texts: list[str] = []
        total_rows = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[str] = []
            for row in ws.iter_rows(values_only=True):
                cells: list[str] = []
                for cell in row:
                    if cell is not None:
                        cell_str = str(cell).strip()
                        cell_str = cell_str.replace("\t", " ").replace("\n", " ")
                        cells.append(cell_str)
                    else:
                        cells.append("")
                if any(c for c in cells):
                    rows.append("\t".join(cells))
                    total_rows += 1

            if rows:
                sheet_texts.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))

        full_text = "\n\n".join(sheet_texts)
        wb.close()

        return ParseResult(
            text=full_text,
            file_path=str(path),
            file_type=path.suffix.lstrip(".").lower(),
            metadata={
                "sheet_count": len(wb.sheetnames),
                "sheet_names": wb.sheetnames,
                "total_rows": total_rows,
                "char_count": len(full_text),
            },
        )
